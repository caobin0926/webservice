from openpyxl import load_workbook
from webservice.ws.common import paths
"""
定义用例类
"""


class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.method = None
        self.data = None
        self.expected = None
        self.actual = None
        self.result = None
        self.issql=None


"""
定义读写excel类
"""


class Excel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb[sheet_name]

    # 定义读取Excel方法
    def reads_case(self):
        data_list = []
        for row in range(2, self.sheet.max_row + 1):
            case = Case()
            case.case_id = self.sheet.cell(row, 1).value
            case.title = self.sheet.cell(row, 2).value
            case.url = self.sheet.cell(row, 3).value
            case.method = self.sheet.cell(row, 4).value
            case.data = self.sheet.cell(row, 5).value
            case.expected = self.sheet.cell(row, 6).value
            case.issql = self.sheet.cell(row, 9).value
            data_list.append(case)
        return data_list

    # 定义写Excel方法
    def write_case(self, case_id, actual, result):
        self.sheet.cell(int(case_id) + 1, 7, value=actual)
        self.sheet.cell(int(case_id) + 1, 8, value=result)
        self.wb.save(self.filename)

    # 定义关闭excel
    def close_excel(self):
        self.wb.close()


if __name__=='__main__':
    ex=Excel(paths.case_dir,'sendMCode')
    re=ex.reads_case()
    ex.write_case(1,2,3)
